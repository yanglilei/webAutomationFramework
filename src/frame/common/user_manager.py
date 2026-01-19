import logging
import threading
from dataclasses import dataclass
from typing import List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class UserInfoLocation:
    # excel文件路径
    workbook_addr: str
    # 工作簿名称
    sheet_name: str
    # 用户名开始单元格
    username_start_cell: str
    # 用户名结束单元格
    username_end_cell: str
    # 密码开始单元格
    password_start_cell: str
    # 密码结束单元格
    password_end_cell: str


class UserManager:
    """
    用户信息操作类
    """
    lock = threading.RLock()
    default_load_counts = 10
    realname_cell_to_username_offset = 2
    remark_cell_to_username_offset = 4
    login_error_desc_cell_to_username_offset = 3
    subject_cell_to_username_offset = 5

    def __init__(self, user_info_location: UserInfoLocation):
        """
        初始化
        :param user_info_location: 用户信息位置
        """
        self.user_info_location = user_info_location
        self.workbook_addr = self.user_info_location.workbook_addr
        self.sheet_name = self.user_info_location.sheet_name
        # self.load_counts = self.user_info_location.load_counts
        self.username_start_cell = self.user_info_location.username_start_cell
        self.username_end_cell = self.user_info_location.username_end_cell
        self.password_start_cell = self.user_info_location.password_start_cell
        self.password_end_cell = self.user_info_location.password_end_cell

    def batch_update_learning_status(self, user_update_infos: List[dict]):
        self.lock.acquire()
        ret = True
        error_msg_prefix = "批量更新用户学习状态失败"
        workbook: Workbook = None
        try:
            workbook: Workbook = load_workbook(filename=self.workbook_addr)
            worksheet: Worksheet = workbook[self.sheet_name]
            if worksheet is not None:
                if user_update_infos is not None and len(user_update_infos) > 0:
                    for username, learning_status in user_update_infos:
                        self._update_val_by_username(workbook, worksheet, username, learning_status,
                                                     self.remark_cell_to_username_offset)
            else:
                # 找不到工作簿的情况
                logging.error("%s，文档【%s】中找不到工作簿【%s】" % (error_msg_prefix, self.workbook_addr, self.sheet_name))
                ret = False
        except PermissionError:
            logging.error("%s，文档【%s】保存失败，请关闭文档！" % (error_msg_prefix, self.workbook_addr))
            ret = False
        except Exception as e:
            logging.error("%s，文档【%s】保存失败：" % (error_msg_prefix, self.workbook_addr), exc_info=True)
            ret = False
        finally:
            if workbook is not None:
                try:
                    workbook.save(self.workbook_addr)
                    workbook.close()
                except:
                    ret = False
                    logging.error("用户文件未关闭，无法更新")
            self.lock.release()
        return ret

    def update_login_msg_by_username(self, username, login_error_desc: str, is_append_update_info=True):
        """
        更新用户登录信息
        :param username:
        :param login_error_desc:
        :return:
        """
        return self._update_val_by_username_template(username, login_error_desc,
                                                     self.login_error_desc_cell_to_username_offset,
                                                     is_append_update_info)

    def update_subject_by_username(self, username, subject_str: str):
        """
        更新用户登录信息
        :param username:
        :param subject_str:
        :return:
        """
        return self._update_val_by_username_template(username, subject_str,
                                                     self.subject_cell_to_username_offset, is_append_update_info=False)

    def update_record_by_username(self, username, field_dict: dict, is_append_update_info=False):
        """
        根据用户名更新记录
        :param username: 用户名
        :param field_dict: 要更新的字段字典；格式{相对记录用户名的位置:值}，例如{2:"dsf", 3:"第三方"}
        :return:
        """
        self.lock.acquire()
        ret = False
        error_msg_prefix = "用户【%s】更新文档失败" % username
        workbook: Workbook = None
        try:
            workbook: Workbook = load_workbook(filename=self.workbook_addr)
            worksheet: Worksheet = workbook[self.sheet_name]
            if worksheet is not None:
                # 判断表中的姓名是否已存在
                username_cell = self._locate_username_cell(username, worksheet)

                for k, v in field_dict.items():
                    # 没有姓名，更新姓名
                    ret = self._update_val_by_username(workbook, worksheet, username, v, k, is_append_update_info)
            else:
                # 找不到工作簿的情况
                logging.error("%s，文档【%s】中找不到工作簿【%s】" % (error_msg_prefix, self.workbook_addr, self.sheet_name))
                ret = False
        except PermissionError:
            logging.error("%s，文档【%s】保存失败，请关闭文档！" % (error_msg_prefix, self.workbook_addr))
            ret = False
        except Exception as e:
            logging.error("%s，文档【%s】保存失败：" % (error_msg_prefix, self.workbook_addr), exc_info=True)
            ret = False
        finally:
            if workbook is not None:
                try:
                    workbook.save(self.workbook_addr)
                    workbook.close()
                except:
                    ret = False
                    logging.error("用户文件未关闭，无法更新")
            self.lock.release()

        return ret

    def update_user_realname_by_username(self, username, realname):
        """
        更新用户姓名
        :param username:
        :param realname:
        :return:
        """
        self.lock.acquire()
        ret = False
        error_msg_prefix = "用户【%s】更新文档失败" % username
        workbook: Workbook = None
        try:
            workbook: Workbook = load_workbook(filename=self.workbook_addr)
            worksheet: Worksheet = workbook[self.sheet_name]
            if worksheet is not None:
                # 判断表中的姓名是否已存在
                username_cell = self._locate_username_cell(username, worksheet)
                realname_cell: Cell = username_cell.offset(0, self.realname_cell_to_username_offset)
                if realname_cell.value is None or len(str(realname_cell.value).strip()) == 0:
                    # 没有姓名，更新姓名
                    ret = self._update_val_by_username(workbook, worksheet, username, realname,
                                                       self.realname_cell_to_username_offset)
            else:
                # 找不到工作簿的情况
                logging.error("%s，文档【%s】中找不到工作簿【%s】" % (error_msg_prefix, self.workbook_addr, self.sheet_name))
                ret = False
        except PermissionError:
            logging.error("%s，文档【%s】保存失败，请关闭文档！" % (error_msg_prefix, self.workbook_addr))
            ret = False
        except Exception as e:
            logging.error("%s，文档【%s】保存失败：" % (error_msg_prefix, self.workbook_addr), exc_info=True)
            ret = False
        finally:
            if workbook is not None:
                try:
                    workbook.save(self.workbook_addr)
                    workbook.close()
                except:
                    ret = False
                    logging.error("用户文件未关闭，无法更新")
            self.lock.release()

        return ret

    def update_learning_status(self, username: str, status_des: str, is_append_update_info=True) -> bool:
        """
        根据用户名更新学习状态
        :param username:用户名
        :param status_des:状态描述
        :return:True-更新成功；False-更新失败
        """
        return self._update_val_by_username_template(username, status_des, self.remark_cell_to_username_offset,
                                                     is_append_update_info)

    def _update_val_by_username_template(self, username, update_info, cell_offset_by_username,
                                         is_append_update_info=True):
        self.lock.acquire()
        ret = False
        error_msg_prefix = "用户【%s】更新文档失败" % username
        workbook: Workbook = None
        try:
            workbook: Workbook = load_workbook(filename=self.workbook_addr)
            worksheet: Worksheet = workbook[self.sheet_name]
            if worksheet is not None:
                ret = self._update_val_by_username(workbook, worksheet, username, update_info, cell_offset_by_username,
                                                   is_append_update_info)
            else:
                # 找不到工作簿的情况
                logging.error("%s，文档【%s】中找不到工作簿【%s】" % (error_msg_prefix, self.workbook_addr, self.sheet_name))
                ret = False
        except PermissionError:
            logging.error("%s，文档【%s】保存失败，请关闭文档！" % (error_msg_prefix, self.workbook_addr), exc_info=True)
            ret = False
        except Exception as e:
            logging.error("%s，文档【%s】保存失败：" % (error_msg_prefix, self.workbook_addr), exc_info=True)
            ret = False
        finally:
            if workbook is not None:
                try:
                    workbook.save(self.workbook_addr)
                    workbook.close()
                except:
                    ret = False
                    logging.error("用户文件未关闭，无法更新")
            self.lock.release()
        return ret

    def _update_val_by_username(self, workbook: Workbook, worksheet: Worksheet, username, update_info,
                                cell_offset_by_username, is_append_update_info=True):
        ret = False
        username_cell = self._locate_username_cell(username, worksheet)
        if username_cell is not None:
            update_info_cell: Cell = username_cell.offset(0, cell_offset_by_username)
            if is_append_update_info:
                content = ("%s;%s" % (
                    update_info_cell.value, update_info)) if update_info_cell.value is not None else update_info
            else:
                content = update_info

            # logging.info("用户【%s】更新单元格的信息：%s" % (username, content))
            worksheet.cell(update_info_cell.row, update_info_cell.column, content)
            ret = True
        else:
            # 用户找不到的情况
            logging.info("用户【%s】找不到，请检查用户名，或者原表是否有改动" % username)
            ret = False
        return ret

    def get_cell_val(self, username, offset_by_username: int):
        self.lock.acquire()
        ret = None
        try:
            workbook: Workbook = load_workbook(filename=self.workbook_addr)
            worksheet: Worksheet = workbook[self.sheet_name]
            username_cell = self._locate_username_cell(username, worksheet)
            target_cell: Cell = username_cell.offset(0, offset_by_username)
            ret = target_cell.value
        finally:
            if not workbook:
                workbook.close()
            self.lock.release()
        return ret

    def get_users(self) -> List[Tuple[str, str]]:
        ret = list()
        login_infos = self._get_login_info_cell()
        for info in login_infos:
            username_cell = info[0]
            password_cell = info[1]
            if username_cell[0].value is None or len(str(username_cell[0].value).strip()) == 0:
                continue
            else:
                ret.append((str(username_cell[0].value).strip() if username_cell[0].value is not None else "",
                            str(password_cell[0].value).strip() if password_cell[0].value is not None else ""))
        return ret

    def _get_login_info_cell(self) -> List[Tuple[Cell, Cell]]:
        """
        获取用户信息
        :return:
        """
        workbook: Workbook = None
        try:
            workbook = load_workbook(filename=self.workbook_addr)
            worksheet: Worksheet = workbook[self.sheet_name]
            usernames: List[Tuple[Cell]] = worksheet["%s:%s" % (self.username_start_cell, self.username_end_cell)]
            passwords: List[Tuple[Cell]] = worksheet["%s:%s" % (self.password_start_cell, self.password_end_cell)]
            if len(usernames) != len(passwords):
                # 用户名和密码个数不匹配
                raise ValueError("用户名和密码个数不匹配")
            return list(zip(usernames, passwords))
        except Exception as e:
            logging.error("获取用户登录信息失败", exc_info=True)
            raise e
        finally:
            if workbook:
                try:
                    workbook.close()
                except:
                    pass

    def _locate_username_cell(self, username, worksheet) -> Cell:
        """
        定位用户名所在的单元格
        :param username:
        :return:Cell 单元格位置
        """
        ret = None
        usernames: List[Tuple[Cell]] = worksheet["%s:%s" % (self.username_start_cell, self.username_end_cell)]
        if usernames is not None and len(usernames) > 0:
            for u in usernames:
                if str(u[0].value).strip() == str(username):
                    ret = u[0]
                    break
        return ret
