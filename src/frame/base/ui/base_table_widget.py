"""
基础表格组件
"""
import functools
import os
from abc import abstractmethod
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, List, Tuple, Any, Optional, Callable

import numpy as np
import pandas as pd
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QIntValidator, QFontMetrics, QColor
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                             QLineEdit, QPushButton, QComboBox, QTableWidget, QTableWidgetItem,
                             QMessageBox, QCalendarWidget,
                             QProgressDialog, QMenu, QCheckBox, QHeaderView, QDialog, QFormLayout, QFileDialog,
                             QListWidget, QListWidgetItem, QTextEdit)
from openpyxl.reader.excel import load_workbook

from src.frame.common.constants import Constants
from src.frame.common.sys_config import SysConfig
from src.frame.common.ui import ShadowButton
from src.frame.dao.async_db_task_scheduler import AsyncTaskScheduler
from src.ui.ui_cell_content_dialog import CellContentDialog


# class BulkImportWorker(QThread):
#     # 批量导入工作线程
#     progress = pyqtSignal(int, int)  # 当前行/总行数
#     finished = pyqtSignal(int)  # 成功条数
#     error = pyqtSignal(str, int)  # 错误信息/行号
#
#     def __init__(self, file_path, batch_size=1000):
#         super().__init__()
#         self.file_path = file_path
#         self.batch_size = batch_size
#         self._cancel_flag = False
#         self.total_rows = 0
#
#     def cancel(self):
#         self._cancel_flag = True
#
#     def run(self):
#         try:
#             # 使用内存映射模式读取Excel
#             wb = load_workbook(
#                 filename=self.file_path,
#                 read_only=True,  # 只读模式
#                 data_only=True,  # 忽略公式
#                 keep_links=False  # 禁用外部链接
#             )
#             ws = wb.active
#
#             # 预扫描总行数（不加载数据）
#             self.total_rows = ws.max_row - 1  # 排除标题行
#             self.progress.emit(0, self.total_rows)
#
#             # 批量插入缓冲区
#             buffer = []
#             success_count = 0
#             current_row = 0
#
#             # 流式逐行读取
#             for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
#                 if self._cancel_flag:
#                     break
#
#                 try:
#                     # 转换为字典（按列名映射）
#                     data = self.parse_row(row)
#                     if self.validate_row(data):
#                         buffer.append(data)
#                         success_count += 1
#
#                     # 批量提交
#                     if len(buffer) >= self.batch_size:
#                         self.bulk_insert(buffer)
#                         buffer = []
#
#                     current_row = idx - 1
#                     self.progress.emit(current_row, self.total_rows)
#
#                 except Exception as e:
#                     self.error.emit(str(e), idx)
#
#             # 提交剩余数据
#             if buffer:
#                 self.bulk_insert(buffer)
#
#             self.finished.emit(success_count)
#
#         except Exception as e:
#             self.error.emit(f"文件读取失败: {str(e)}", 0)
#         finally:
#             if 'wb' in locals():
#                 wb.close()
#
#     def parse_row(self, row):
#         """将Excel行解析为字典（需子类实现）"""
#         return {
#             'col1': row[0].value,
#             'col2': row[1].value
#         }
#
#     def validate_row(self, data):
#         """行数据验证（需子类实现）"""
#         return True
#
#     def bulk_insert(self, data_list):
#         """批量插入方法（需子类实现）"""
#         pass


@dataclass
class EditableField:
    field_name: str  # 字段名
    # 字段类型
    # 可选值如下：
    # text：可编辑文本
    # select：单选
    # date：日期
    # label: 标签
    # textedit: 多行文本
    field_type: str = 'text'
    select_options: list = None  # select类型的可选值，第一个元素为名称，第二个元素为对应的值，例如：[(name1, value1), (name2, value2)]
    related_field: Optional[str] = None  # 关联字段名（如project_name）
    # 关联值获取器：接收当前选中的选项索引，返回关联字段的值
    # TODO 该字段的用法说明需要完善！！！pause by zcy 20260117
    related_value_getter: Optional[Callable[[int], Any]] = None
    # 生成的组件
    widget: QWidget = None
    # 字段属性
    # 隐藏/显示
    visible: bool = True
    # 只读
    readonly: bool = False

    def __post_init__(self):
        if self.select_options is None:
            self.select_options = []


@dataclass
class TableHeader:
    # 表头显示名称
    show_name: str
    # 数据字段名，对应数据源的字段名
    field_name: str
    # 是否在表格中显示
    is_column_visible: bool = True
    # 是否可添加。默认为True，标记为false时，不会将该字段保存到表格中
    is_add_visible: bool = True
    # 是否可编辑。默认为True，标记为false时，不会将该字段保存到表格中
    is_edit_visible: bool = True


@dataclass
class QueryField:
    # 组件类型
    widget_type: str
    # 标签
    label: str
    # 数据字段名
    field_name: str
    # QComboBox的选项
    options: List = field(default_factory=list)
    # 默认值
    default_value: Any = None
    # 占位符
    placeholder: str = ''
    # 异步函数，当widget_type=select时，需要异步获取数据时，提供该回调函数，该函数返回一个列表List[tuple[str, Any]]，每个元素为：(中文名, data)
    async_func: Optional[Callable] = None
    # 控制组件的最小宽度
    min_width: Optional[int] = None


class ExportWorker(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, data, file_path, columns):
        super().__init__()
        self.data = data
        self.file_path = file_path
        self.columns = columns

    def run(self):
        try:
            # 转换数据并筛选字段
            df = pd.DataFrame(self.data)
            df = df[self.columns.keys()]
            df.rename(columns=self.columns, inplace=True)

            # 分块写入（适用于大数据量）
            data_length = len(self.data)
            if data_length <= 10000:
                with pd.ExcelWriter(self.file_path, engine='xlsxwriter') as writer:  # 使用上下文管理器
                    # df.to_excel(writer, chu=1000)  # 分块写入
                    df.to_excel(writer, index=False)
                    self.progress.emit(data_length)
            else:
                with pd.ExcelWriter(self.file_path, engine='xlsxwriter') as writer:
                    # 转换为DataFrame（如果数据是原始字典列表）
                    total_rows = len(df)
                    # 分块处理（每1000行）
                    chunk_size = 1000
                    for start in range(0, total_rows, chunk_size):
                        end = min(start + chunk_size, total_rows)
                        chunk = df.iloc[start:end]

                        # 写入Excel（动态处理表头）
                        header = (start == 0)  # 仅第一块写表头
                        chunk.to_excel(
                            writer,
                            # sheet_name="Data",
                            startrow=start if header else start + 1,
                            header=header,
                            index=False
                        )
                        self.progress.emit(end)  # 更新进度

            self.finished.emit(self.file_path)
        except Exception as e:
            self.error.emit(str(e))


class FieldSelectDialog(QDialog):
    def __init__(self, headers: List[TableHeader]):
        super().__init__()
        self.setWindowTitle("选择导出字段")
        self.selected_fields = []

        layout = QVBoxLayout()
        self.list_widget = QListWidget()

        # 生成带复选框的字段列表
        for header in headers:  # 跳过复选框列
            if not header.is_column_visible:
                continue
            item = QListWidgetItem()
            cb = QCheckBox(header.show_name)
            cb.setChecked(True)
            cb.field_name = header.field_name
            self.list_widget.addItem(item)
            self.list_widget.setItemWidget(item, cb)

        # 操作按钮
        btn_select_all = QPushButton("全选")
        btn_select_all.clicked.connect(lambda: self.toggle_all(True))
        btn_unselect_all = QPushButton("全不选")
        btn_unselect_all.clicked.connect(lambda: self.toggle_all(False))
        btn_confirm = QPushButton("确定")
        btn_confirm.clicked.connect(self.accept)

        layout.addWidget(self.list_widget)
        layout.addWidget(btn_select_all)
        layout.addWidget(btn_unselect_all)
        layout.addWidget(btn_confirm)
        self.setLayout(layout)

    def toggle_all(self, checked):
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            widget = self.list_widget.itemWidget(item)
            widget.setChecked(checked)

    def get_selected_fields(self) -> Dict[str, str]:
        """返回选择的字段映射 {字段名: 显示文本}"""
        selected = {}
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            widget = self.list_widget.itemWidget(item)
            if widget.isChecked():
                selected[widget.field_name] = widget.text()
        return selected


class BaseAsyncImportWorker(QThread):
    progress = pyqtSignal(int, int)  # 当前进度/总条数
    finished = pyqtSignal(int)  # 成功导入数
    error = pyqtSignal(str, int)  # 错误信息/出错行号

    def __init__(self, file_path="", headers: Optional[Dict] = None, batch_size=1000):
        """
        参数说明：
        - file_path: Excel文件路径
        - headers: 中英文字段映射字典，格式：{'中文列名': 'field_name', ...}
        - batch_size: 批量提交大小
        """
        super().__init__()
        self.file_path = file_path
        self.headers = headers
        self.batch_size = batch_size
        self._cancel_flag = False
        self.total_rows = 0
        self.column_map = {}  # 存储Excel列位置与字段名的映射

    def cancel(self):
        self._cancel_flag = True

    def run(self):
        wb = None
        try:
            # 初始化Excel读取
            wb = load_workbook(
                filename=self.file_path,
                read_only=True,
                data_only=True,
                keep_links=False
            )
            ws = wb.active

            # 解析表头（自动识别中文列名位置）
            self._parse_header(ws[1])  # 假设标题在第一行

            # 计算总行数（排除标题行）
            self.total_rows = ws.max_row - 1
            self.progress.emit(0, self.total_rows)

            # 初始化批量缓冲区
            buffer = []
            success_count = 0
            current_row = 0

            # 流式处理数据行
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if self._cancel_flag:
                    break

                try:
                    # 自动解析行数据
                    data = self._parse_row(row)
                    if self.validate_row(data):
                        buffer.append(data)
                        success_count += 1

                    # 批量提交
                    if len(buffer) >= self.batch_size:
                        self.bulk_insert(buffer)
                        buffer = []

                    # 更新进度
                    current_row = idx - 2  # 标题行偏移
                    self.progress.emit(current_row, self.total_rows)

                except Exception as e:
                    self.error.emit(f"行处理失败: {str(e)}", idx)

            # 提交剩余数据
            if buffer:
                self.bulk_insert(buffer)

            self.finished.emit(success_count)

        except Exception as e:
            self.error.emit(f"文件处理失败: {str(e)}", 0)
        finally:
            if wb:
                wb.close()

    def _parse_header(self, header_row):
        """解析Excel表头行，建立列位置到字段名的映射"""
        self.column_map.clear()

        for idx, cell in enumerate(header_row):
            chinese_name = cell.value
            if chinese_name in self.headers:
                field_name = self.headers[chinese_name]
                self.column_map[idx] = field_name

    def _parse_row(self, row):
        """自动将Excel行转换为字段字典"""
        row_data = {}
        for idx, cell in enumerate(row):
            if idx in self.column_map:
                field_name = self.column_map[idx]
                row_data[field_name] = cell.value
        return row_data

    @abstractmethod
    def bulk_insert(self, data_list) -> Tuple[bool, str]:
        """执行批量插入"""
        pass
        # try:
        #     # 通过table_widget访问DAO
        #     self.table_widget.dao.batch_insert_image_info(data_list)
        # except Exception as e:
        #     error_msg = f"批量插入失败: {str(e)}"
        #     self.error.emit(error_msg, 0)
        #     raise  # 向上抛出以停止导入流程

    def validate_row(self, data):
        """基础数据验证（可扩展）"""
        return True


class BaseTableWidget(QWidget):
    """
    基础表格组件基类
    支持翻页，导入、导出，条件筛选
    """
    # 定义更新结果信号（异步操作数据时使用）
    update_finished = pyqtSignal(bool, str, object)  # 成功标志、消息、额外数据
    # 添加一条记录成功的信号（异步操作数据时使用）
    add_one_finished = pyqtSignal(bool, str, object)  # 成功标志、消息、额外数据
    # 删除行成功的信号（异步操作数据时使用）
    delete_finished = pyqtSignal(bool, str, object)  # 成功标志、消息、额外数据
    # 清空所有数据成功信号（异步操作数据时使用）
    clear_all_finished = pyqtSignal(bool, str, object)

    def __init__(self, is_support_import=False, is_need_search=False, is_support_add=True, is_support_clear_all=False,
                 is_support_export=False, is_support_open_context_menu=True):
        super().__init__()
        ###### 定义属性 ######
        self.bulk_importer = None  # 批量导入处理
        self.import_progress_dialog = None  # 导入进度对话框
        self.add_loading_dialog = None  # 新增记录的加载对话框
        self.edit_loading_dialog = None  # 进入编辑页面的加载对话框
        self.export_selected_fields: Optional[Dict[str, str]] = None  # 导出字段
        self.export_file_path: Optional[str] = None  # 导出文件路径
        self.export_progress_dialog = None  # 导出进度对话框
        self.export_thread = None  # 导出线程
        self.edit_fields: Dict[str, EditableField] = {}  # 可编辑字段
        self.add_fields: Dict[str, EditableField] = {}  # 新增的字段
        self.cmb_page_size = None  # 页大小下拉框
        self.btn_go = None  # 跳转按钮
        self.le_page_num = None  # 页码输入框
        self.btn_next = None  # 下一页按钮
        self.lbl_page = None  # 页码标签
        self.lbl_total = None  # 总页数标签
        self.btn_prev = None  # 上一页按钮
        self.is_support_import = is_support_import  # 是否支持导入
        self.is_need_search = is_need_search  # 是否需要搜索
        self.is_support_add = is_support_add  # 是否支持新增
        self.is_support_clear_all = is_support_clear_all  # 是否支持清空所有数据
        self.is_support_open_context_menu = is_support_open_context_menu  # 是否支持右键菜单
        self.is_support_export = is_support_export
        self.page_num = 1  # 页码
        self.page_size = 10  # 每页条数
        self.total_pages = 0  # 总页数
        self.selected_rows = set()  # 选中的行ID
        self.field_map: Optional[Dict[str, str]] = None  # 字段映射，key-表头名称；value-数据库的字段名称
        self.edit_dialog: Optional[QDialog] = None  # 编辑对话框
        self.add_one_dialog: Optional[QDialog] = None  # 新增以条记录的对话框
        self.search_progress_dialog: Optional[QProgressDialog] = None  # 搜索进度对话框
        self.query_widgets = {}  # 搜索条件
        self.main_layout = QVBoxLayout()  # 主布局
        self.table = QTableWidget()  # 表格控件

        ###### 初始化UI ######
        self.init_ui()  # 初始化UI

        ###### 异步加载数据 ######
        self.current_page_rows = set()  # 新增：记录当前页行号
        self.async_task_scheduler = AsyncTaskScheduler()  # 异步DB任务调度器
        self.first_load_data()  # 第一次加载数据

        ###### 监听信号 ######
        self.update_finished.connect(self.on_update_result)
        self.add_one_finished.connect(self.on_add_one_result)
        self.delete_finished.connect(self.on_delete_result)
        self.clear_all_finished.connect(self.on_delete_result)

    def first_load_data(self):
        self.async_get_data(self.first_load_success, True)  # 异步刷新表格数据

    def first_load_success(self, status: bool, msg: str, payloads: object):
        self.on_first_load_success(payloads)
        self.render_table(status, msg, payloads)
        self.after_first_render_table(payloads)

    def on_first_load_success(self, payloads: object):
        """
        初始化数据成功后，渲染表格之前的回调
        :param payloads: get_records方法的返回值
        :return:
        """
        pass

    def after_first_render_table(self, payloads):
        """
        初次渲染表格成功后的回调
        :param payloads: get_records方法的返回值
        :return:
        """
        pass

    def show_add_dialog(self):
        """重构后的新增方法：先请求最新元数据，再渲染弹窗"""
        # 1. 显示加载提示
        # self.add_loading_dialog = QProgressDialog("正在加载最新数据...", "取消", 0, 0, self)
        # self.add_loading_dialog.setWindowModality(Qt.WindowModal)
        # self.add_loading_dialog.show()

        # 2. 异步请求最新的新增元数据
        self.async_task_scheduler.submit_task(
            self.get_add_metadata,  # 子类实现：获取新增所需的最新元数据
            finished_callback=self.on_add_metadata_loaded
        )

    def on_add_metadata_loaded(self, status: bool, msg: str, metadata: dict):
        """新增元数据加载完成后的回调"""
        # 关闭加载提示
        # if self.add_loading_dialog and self.add_loading_dialog.isVisible():
        #     self.add_loading_dialog.close()

        if not status:
            QMessageBox.warning(self, "错误", f"加载最新数据失败：{msg}")
            return

        # 渲染新增弹窗（使用最新元数据）
        self.render_add_dialog(metadata)

    def render_add_dialog(self, metadata: dict):
        """渲染新增弹窗（使用最新元数据）"""
        self.add_one_dialog = QDialog(self)
        self.add_one_dialog.setWindowTitle("新增记录")
        layout = QVBoxLayout(self.add_one_dialog)

        # 动态生成表单字段
        self.add_fields: Dict[str, Any] = {}
        form_layout = QFormLayout()
        headers = self.get_headers()[1:]  # 跳过ID列

        for header in headers:
            if not header.is_add_visible:
                continue
            field_name = header.field_name
            show_name = header.show_name

            # 根据最新元数据创建组件
            field = self._create_add_widget_with_metadata(field_name, "", metadata)
            if field and field.widget:
                if not field.visible:
                    continue
                widget = field.widget
                if hasattr(widget, "setVisible"):
                    widget.setVisible(field.visible)
                if hasattr(widget, "setReadOnly"):
                    widget.setReadOnly(field.readonly)
                self.add_fields[field_name] = field
                form_layout.addRow(QLabel(show_name + ":"), field.widget)

        # 按钮布局（原有逻辑）
        btn_submit = QPushButton("提交")
        btn_submit.clicked.connect(self.save_add_record)
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.add_one_dialog.reject)

        layout.addLayout(form_layout)
        layout.addWidget(btn_submit)
        layout.addWidget(btn_cancel)
        self.add_one_dialog.exec_()

    def _create_add_widget_with_metadata(self, field_name: str, value: str, metadata: dict) -> Optional[EditableField]:
        """根据字段类型+最新元数据创建新增组件"""
        # 从元数据中获取该字段的最新属性
        field_attr = metadata.get(field_name, EditableField(field_name, "text", QLineEdit()))

        field_type = field_attr.field_type
        if field_type == "select":
            widget = QComboBox()
            options = field_attr.select_options  # 最新的下拉选项
            for option in options:
                widget.addItem(option[0], option[1])
            current_index = next((i for i, (k, _) in enumerate(options) if k == value), 0)
            widget.setCurrentIndex(current_index)
        elif field_type == "date":
            widget = QLineEdit(value)
            widget.setPlaceholderText("YYYY-MM-DD")
            widget.mousePressEvent = self.create_date_handler(widget)
        elif field_type in "hide":
            widget = None
        elif field_type == "textedit":
            widget = QTextEdit(value)
        else:
            widget = QLineEdit(value)

        field_attr.widget = widget
        return field_attr

    def save_add_record(self):
        """保存新记录"""
        data = {}
        for field_name, field in self.add_fields.items():
            widget = field.widget
            if isinstance(widget, QComboBox):
                if not field.related_field:
                    data[field_name] = widget.currentData()
                else:
                    data[field_name] = widget.currentText()
                    data[field.related_field] = widget.currentData()
            elif isinstance(widget, QLineEdit):
                data[field_name] = widget.text()
            elif isinstance(widget, QTextEdit):
                data[field_name] = widget.toPlainText()
            elif isinstance(widget, (str, int, float)):
                # 添加自定义字段值的时候用到
                data[field_name] = widget

        # 合并隐藏字段（原有逻辑）
        # hidden_fields = self.add_hidden_fields()
        # data.update(hidden_fields)

        if self.prepare_for_add(data):
            self.do_add_one(data)

    def async_import(self):
        # 禁用按钮防止重复点击
        self.btn_import.setEnabled(False)
        path_obj = SysConfig.get_value(Constants.ConfigFileKey.LATEST_DATA_FILE_DIR_NAME)
        if not path_obj or not path_obj.get("value"):
            path = os.getcwd()
        else:
            path = path_obj.get("value")
        file_path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", path, "Excel Files (*.xlsx *.xls)")
        if not file_path:
            self.btn_import.setEnabled(True)
            return

        SysConfig.save_value(Constants.ConfigFileKey.LATEST_DATA_FILE_DIR_NAME, os.path.dirname(path))
        self.le_data_file_path.setText(file_path)
        # 初始化进度条
        self.import_progress_dialog = QProgressDialog("正在导入...", "取消", 0, 100, self)
        self.import_progress_dialog.show()
        self.import_progress_dialog.canceled.connect(self.cancel_import)
        self.bulk_importer = self.create_bulk_importer()  # 异步导入线程类
        self.bulk_importer.file_path = file_path
        # TODO 需要修改！因为表头的顺序、内容和数据库的不一致，所以需要修改
        # self.bulk_importer.headers = {item[0]: item[1] for item in self.get_headers()}
        self.bulk_importer.batch_size = 2000

        # 启动线程
        self.bulk_importer.progress.connect(self.update_import_progress)
        self.bulk_importer.finished.connect(self.import_completed)
        self.bulk_importer.error.connect(self.handle_import_error)
        self.bulk_importer.start()

    def cancel_import(self):
        self.bulk_importer.cancel()
        self.btn_import.setEnabled(True)

    def update_import_progress(self, current, total):
        """更新进度条"""
        self.import_progress_dialog.setValue(int(current / total * 100))

    def import_completed(self, count):
        self.import_progress_dialog.close()
        self.btn_import.setEnabled(True)
        QMessageBox.information(self, "完成", f"成功导入 {count} 条记录")
        self.async_refresh_table()

    def handle_import_error(self, msg, row):
        self.import_progress_dialog.setLabelText(f"第{row}行错误：{msg}")

    def import_from_excel(self):
        """从Excel导入数据"""
        path_obj = SysConfig.get_value(Constants.ConfigFileKey.LATEST_DATA_FILE_DIR_NAME)
        if not path_obj or not path_obj.get("value"):
            path = os.getcwd()
        else:
            path = path_obj.get("value")
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", path, "Excel Files (*.xlsx *.xls)"
        )
        if not file_path:
            return

        SysConfig.save_value(Constants.ConfigFileKey.LATEST_DATA_FILE_DIR_NAME, os.path.dirname(path))

        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            # 1. 处理NaN值
            df = df.replace({np.nan: None})  # 将NaN统一转为Python的None
            # 转换列名为字段名
            column_mapping = {h.show_name: h.field_name for h in self.get_headers()[1:]}
            df.rename(columns=column_mapping, inplace=True)

            # 验证数据
            valid_data = []
            for _, row in df.iterrows():
                if self.validate_import_data(row.to_dict()):
                    valid_data.append(row.to_dict())
        except:
            QMessageBox.warning(self, "错误", "导入数据时发生错误")
        else:
            # 批量插入
            if valid_data:
                self.do_batch_insert(valid_data)

    def export_to_excel(self):
        """优化后的导出方法"""
        # 1. 字段选择
        dialog = FieldSelectDialog(self.get_headers())
        if dialog.exec_() != QDialog.Accepted:
            return

        self.export_selected_fields = dialog.get_selected_fields()
        if not self.export_selected_fields:
            QMessageBox.warning(self, "警告", "请至少选择一个导出字段")
            return

        # 2. 文件路径选择
        self.export_file_path, _ = QFileDialog.getSaveFileName(
            self, "保存Excel文件", "", "Excel Files (*.xlsx)"
        )
        if not self.export_file_path:
            return

        # 3. 获取数据（异步）
        self.export_progress_dialog = QProgressDialog("正在准备数据...", "取消", 0, 0, self)
        self.export_progress_dialog.show()
        self.export_progress_dialog.setWindowModality(Qt.WindowModal)  # 阻塞父窗口但不冻结UI
        self.async_task_scheduler.submit_task(
            self.get_records, self.build_query_condition(), 1, 0,
            finished_callback=self.start_export)

        # 启动后台线程获取数据
        # self.async_get_data(lambda data: self.start_export(data, file_path, selected_fields))
        # self.data_fetcher = (self, 1, 0)
        # self.data_fetcher.finished.connect(
        #     lambda data: self.start_export(data, file_path, selected_fields))
        # self.data_fetcher.start()

    def start_export(self, status: bool, msg: str, data):
        """启动导出线程"""
        # 关闭数据获取进度
        self.export_progress_dialog.close()
        if not status:
            QMessageBox.warning(self, "错误", f"导出失败：{msg}")
            return

        # 配置导出进度
        self.export_progress_dialog = QProgressDialog(
            "正在导出数据...", "取消", 0, data[1], self)
        self.export_progress_dialog.show()
        self.export_progress_dialog.setWindowModality(Qt.WindowModal)  # 阻塞父窗口但不冻结UI

        # 创建导出线程
        self.export_thread = ExportWorker(data[0], self.export_file_path, self.export_selected_fields)
        self.export_thread.progress.connect(self.update_progress)
        self.export_thread.finished.connect(self.export_success)
        self.export_thread.error.connect(self.export_failed)
        self.export_thread.start()

    def update_progress(self, value):
        self.export_progress_dialog.setValue(value)

    def export_success(self, file_path):
        self.export_progress_dialog.close()
        QMessageBox.information(self, "导出成功", f"文件已保存至：\n{file_path}")

    def export_failed(self, error_msg):
        self.export_progress_dialog.close()
        QMessageBox.critical(self, "导出失败", f"发生错误：\n{error_msg}")

    def _create_import_layout(self):
        ly_import = QHBoxLayout()
        self.btn_import = QPushButton("Excel导入")
        self.btn_import.clicked.connect(self.async_import)

        self.le_data_file_path = QLineEdit()
        self.le_data_file_path.setPlaceholderText("数据表xlsx")

        ly_import.addWidget(QLabel("数据表："))
        ly_import.addWidget(self.le_data_file_path)
        ly_import.addWidget(self.btn_import)
        return ly_import

    def _init_table_widget(self):
        # 初始化表头
        self.init_table_headers()

        # 设置QTableWidget的属性
        # 核心设置：确保行高可拖拽（默认就是Interactive，可显式声明）
        # self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)

        # 可选：设置行高的最小值（避免拖拽到过窄）
        # self.table.verticalHeader().setMinimumSectionSize(30)  # 最小行高30px
        # 可选：设置行高的最大值（避免拖拽到过宽）
        # self.table.verticalHeader().setMaximumSectionSize(200)  # 最大行高200px

        # self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        # self.table.setSelectionBehavior(QTableWidget.SelectItems)
        # self.table.setSelectionMode(QTableWidget.ExtendedSelection)
        # 自定义的列宽自适应
        # self.set_column_width_adaptive()
        # 多选列处理
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        # self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)

        # 核心：绑定双击单元格信号
        self.table.cellDoubleClicked.connect(self.on_cell_double_clicked)
        # 右键菜单
        if self.is_support_open_context_menu:
            self.table.setContextMenuPolicy(Qt.CustomContextMenu)
            self.table.customContextMenuRequested.connect(self.show_context_menu)

    def set_column_width_adaptive(self):
        """设置列宽：除最后一列外自适应内容，最后一列占剩余空间"""
        table = self.table
        column_count = table.columnCount()
        horizontal_header = table.horizontalHeader()

        # 第一步：让所有列先自适应内容宽度
        table.resizeColumnsToContents()

        # 第二步：遍历前n-1列，保证列宽至少不小于表头宽度
        total_width = 0
        for col in range(column_count - 1):
            # 获取内容自适应后的宽度
            content_width = table.columnWidth(col)
            # 获取表头的最小显示宽度（精确计算表头文字所需宽度）
            header_item = horizontal_header.model().headerData(col, Qt.Horizontal)
            header_width = horizontal_header.fontMetrics().boundingRect(str(header_item)).width() + 20  # +20是预留边距

            # 取两者中的较大值作为列宽
            final_width = max(content_width, header_width)
            table.setColumnWidth(col, final_width)
            total_width += final_width  # 累加前n-1列的总宽度

        # 第三步：计算最后一列的宽度
        # 考虑垂直滚动条宽度，避免计算偏差
        scrollbar_width = table.verticalScrollBar().width() if table.verticalScrollBar().isVisible() else 0
        available_width = table.viewport().width() - scrollbar_width

        # 最后一列宽度 = 剩余空间，且最小宽度设为100
        last_col_width = max(available_width - total_width, 100)
        table.setColumnWidth(column_count - 1, last_col_width)

    def on_cell_double_clicked(self, row, col):
        """双击单元格的处理逻辑"""
        # 1. 容错：获取单元格对象，避免空单元格报错
        cell_item = self.table.item(row, col)
        if not cell_item:
            return

        # 2. 获取单元格完整文本（即使界面显示省略，底层是完整的）
        cell_text = cell_item.text()

        # 3. 弹出完整内容窗口
        dialog = CellContentDialog(cell_text, row, col, self)
        dialog.exec_()  # 显示模态窗口

    def init_ui(self):
        if self.is_support_import:
            self.main_layout.addLayout(self._create_import_layout())  # 导入数据布局

        if self.is_need_search:
            self.main_layout.addLayout(self._create_query_ui())  # 查询条件布局

        self._init_table_widget()  # 初始化表格
        self.main_layout.addWidget(self.table)  # 表格添加到布局
        self.main_layout.addLayout(self.create_toolbar_layout())  # 表格下方的工具栏布局
        self.main_layout.addLayout(self.create_pagination_ui())  # 分页布局

        self.setLayout(self.main_layout)

    def select_all_current_page(self):
        """全选当前页"""
        for row in self.current_page_rows:
            if row not in self.selected_rows:
                checkbox = self.table.cellWidget(row, 0)
                self.get_children(checkbox.layout(), QCheckBox)[0].setChecked(True)
                # checkbox.setChecked(True)
                self.selected_rows.add(row)

    def get_children(self, layout, widget_type):
        # 获取子控件数量
        count = layout.count()
        specified_widgets = []
        for i in range(count):
            # 获取子控件
            item = layout.itemAt(i)
            if item.widget() and isinstance(item.widget(), widget_type):
                specified_widgets.append(item.widget())
        return specified_widgets
        # print(f"找到 {len(specified_widgets)} 个 {widget_type.__name__} 类型的子控件。")
        # for widget in specified_widgets:
        #     print(widget)

    def invert_selection_current_page(self):
        """反选当前页"""
        for row in self.current_page_rows:
            checkbox = self.get_children(self.table.cellWidget(row, 0).layout(), QCheckBox)[0]
            new_state = not checkbox.isChecked()
            checkbox.setChecked(new_state)
            if new_state:
                self.selected_rows.add(row)
            else:
                self.selected_rows.discard(row)

    def update_selection(self, state: int, row: int):
        """更新选中状态时验证是否在当前页"""
        if row in self.current_page_rows:
            if state == Qt.Checked:
                self.selected_rows.add(row)
            else:
                self.selected_rows.discard(row)

    def create_toolbar_layout(self):
        """创建批量操作按钮组"""
        batch_layout = QHBoxLayout()

        # 全选/反选按钮
        btn_select_all = QPushButton("全选")
        btn_select_all.clicked.connect(self.select_all_current_page)
        btn_invert_select = QPushButton("反选")
        btn_invert_select.clicked.connect(self.invert_selection_current_page)
        btn_select_all.setStyleSheet("""
            QPushButton {
                min-width: 80px;
                padding: 5px;
                background: #f0f0f0;
            }
        """)

        btn_invert_select.setStyleSheet("""
            QPushButton {
                min-width: 80px;
                padding: 5px;
                background: #f8f8f8;
            }
        """)
        batch_layout.addWidget(btn_select_all)
        batch_layout.addWidget(btn_invert_select)
        batch_layout.addStretch()

        """初始化操作工具栏"""
        # toolbar = QHBoxLayout()

        # 新增按钮
        btn_add = QPushButton("新增记录")
        btn_add.clicked.connect(self.show_add_dialog)
        if self.is_support_add:
            batch_layout.addWidget(btn_add)

        btn_export = QPushButton("Excel导出")
        btn_export.clicked.connect(self.export_to_excel)
        if self.is_support_export:
            batch_layout.addWidget(btn_export)

        # 批量操作按钮
        btn_batch_delete = QPushButton("批量删除")
        btn_batch_delete.clicked.connect(self.batch_delete)
        batch_layout.addWidget(btn_batch_delete)
        if self.is_support_clear_all:
            # 清空所有
            btn_clear_all = ShadowButton("清空所有")
            btn_clear_all.clicked.connect(self.clear_all)
            batch_layout.addWidget(btn_clear_all)

        for widget in self.add_widget_to_toolbar():
            batch_layout.addWidget(widget)
        return batch_layout

    def add_widget_to_toolbar(self) -> List[QWidget]:
        """
        添加自定义控件到工具栏
        子类可自行扩展
        :return : List[QWidget] 控件列表
        """
        return []

    def _create_query_ui(self):
        """动态创建查询条件UI"""
        query_layout = QHBoxLayout()
        for field_cfg in self.get_query_fields():
            label = QLabel(f"{field_cfg.label}:")
            if field_cfg.widget_type == 'date_range':
                widget_group = QWidget()
                # 日期范围特殊布局
                layout = QHBoxLayout(widget_group)
                layout.setContentsMargins(0, 0, 0, 0)

                # 开始日期
                start_widget = QLineEdit()
                start_widget.setPlaceholderText("开始日期")
                start_widget.mousePressEvent = self.create_date_handler(start_widget)
                layout.addWidget(start_widget)

                # 分隔符
                layout.addWidget(QLabel("~"))

                # 结束日期
                end_widget = QLineEdit()
                end_widget.setPlaceholderText("结束日期")
                end_widget.mousePressEvent = self.create_date_handler(end_widget)
                layout.addWidget(end_widget)

                # 存储组件
                self.query_widgets[f"{field_cfg.field_name}_start"] = start_widget
                self.query_widgets[f"{field_cfg.field_name}_end"] = end_widget

                # 添加组件到布局
                query_layout.addWidget(label)
                query_layout.addWidget(widget_group)
            else:
                widget = None
                if field_cfg.widget_type == 'select':
                    widget = QComboBox()
                    widget.addItem("全部", None)
                    options = field_cfg.options
                    if options:
                        for text, value in options:
                            widget.addItem(text, value)
                    _async = field_cfg.async_func
                    if _async:
                        # 异步加载数据
                        self.async_task_scheduler.submit_task(_async, finished_callback=functools.partial(
                            self._add_combox_items, widget))
                    widget.setCurrentIndex(0)
                    widget.setMinimumWidth(150 if not field_cfg.min_width else field_cfg.min_width)
                    widget.setMinimumHeight(35)
                elif field_cfg.widget_type == 'date':
                    widget = QLineEdit()
                    widget.setPlaceholderText("YYYY-MM-DD")
                    widget.mousePressEvent = self.create_date_handler(widget)
                elif field_cfg.widget_type == 'text':
                    widget = QLineEdit()
                    if field_cfg.placeholder:
                        widget.setPlaceholderText(field_cfg.placeholder)

                if widget:
                    query_layout.addWidget(label)
                    query_layout.addWidget(widget)
                    self.query_widgets[field_cfg.field_name] = widget

        self.btn_query = QPushButton("查询")
        self.btn_query.clicked.connect(lambda: self.search_with_progress("正在查询中..."))
        query_layout.addWidget(self.btn_query)
        return query_layout

    def _add_combox_items(self, widget, status, msg, items: List[tuple[str, Any]]):
        for text, value in items:
            widget.addItem(text, value)

    def create_date_handler(self, widget):
        """创建日期选择处理"""
        calendar = QCalendarWidget()
        calendar.setWindowFlags(Qt.Popup)

        def show_calendar(event):
            pos = widget.mapToGlobal(widget.rect().bottomLeft())
            calendar.move(pos)
            calendar.show()

        def select_date(date):
            widget.setText(date.toString("yyyy-MM-dd"))
            calendar.hide()

        calendar.clicked.connect(select_date)
        return show_calendar

    def init_table_headers(self):
        """初始化表头"""
        headers = [TableHeader('选择', '_checkbox')] + self.get_headers()
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels([h.show_name for h in headers])
        self.field_map = {h.show_name: h.field_name for h in headers}

        # 排除隐藏的表头
        for idx, header in enumerate(headers):
            if not header.is_column_visible:
                self.table.setColumnHidden(idx, True)

    def async_get_data(self, finished_callback: Callable, is_first_load_data=False):
        # self.async_task_scheduler.submit_task(
        #     self.get_records, {} if is_first_load_data else self.build_query_condition(), self.page_num, self.page_size,
        #     finished_callback=finished_callback)
        self.async_task_scheduler.submit_task(
            self.get_records, self.build_query_condition(), self.page_num, self.page_size,
            finished_callback=finished_callback)

    def async_refresh_table(self):
        """异步刷新表格数据"""
        self.async_get_data(self.render_table)

    def render_table(self, status: bool, msg: str, table_data: Any):
        """刷新表格数据"""
        # # 构建查询条件
        # condition = self.build_query_condition()
        # # 获取数据
        # data_list, total = self.get_data(condition, self.page_num, self.page_size)
        if not status:
            QMessageBox.warning(self, "错误", msg)
            return
        data_list, total = table_data
        self.table.setRowCount(len(data_list))
        for row, data in enumerate(data_list):
            # 复选框列
            # chk = QCheckBox()
            # chk.stateChanged.connect(lambda state, r=row: self.update_selection(state, r))
            # self.table.setCellWidget(row, 0, chk)
            check_box = QCheckBox()
            check_box.stateChanged.connect(lambda state, r=row: self.update_selection(state, r))
            check_box.setStyleSheet("margin: 0 auto;")  # 设置水平居中样式
            check_box.setStyleSheet(
                """QCheckBox::indicator { 
                    width: 25px; 
                    height: 25px; 
                } 
                """
            )
            # check_box.setStyleSheet("QWidget { background-color: white; }")
            cell_widget = QWidget()
            layout = QHBoxLayout(cell_widget)
            layout.addWidget(check_box)
            layout.setAlignment(Qt.AlignCenter)  # 整体布局在单元格中居中
            layout.setContentsMargins(0, 0, 0, 0)
            cell_widget.setLayout(layout)
            cell_widget.setStyleSheet("QWidget { background-color: white; }")
            # cell_widget.setBackground(QBrush(QColor(255, 255, 255)))
            self.table.setCellWidget(row, 0, cell_widget)
            # self.table.setItem(row, 0, cell_widget)
            # self.table.item(row, 0).setBackground(QColor(255, 255, 255))

            # 数据列
            for col in range(1, self.table.columnCount()):
                header_text = self.table.horizontalHeaderItem(col).text()
                field_name = self.field_map[header_text]
                value = data.get(field_name, "")
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(value, date):
                    value = value.strftime("%Y-%m-%d")
                elif mapping_val := self.field_mapping(field_name, value):
                    value = mapping_val
                # elif isinstance(value, bool):
                #     value = "是" if value else "否"
                elif value is None:
                    value = ""
                else:
                    value = str(value)
                # value = str(data.get(field_name, ""))
                item = QTableWidgetItem(value)
                # item.setFlags(item.flags() | Qt.ItemIsSelectable)  # 确保可选择
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # 禁用编辑
                self.table.setItem(row, col, item)

        self.total_pages = (total + self.page_size - 1) // self.page_size
        self.lbl_total.setText(f"共{total}条")
        self.lbl_page.setText(f"{self.page_num}/{self.total_pages}")
        self.update_pagination(total)
        # 刷新表格时更新当前页行号记录
        self.current_page_rows = set(range(self.table.rowCount()))
        # 刷新表格时，清空已经选中的行
        self.selected_rows.clear()

    def set_row_background(self, row_index, color: QColor):
        """
        给QTableWidget的指定行设置背景色
        :param table: QTableWidget对象
        :param row_index: 行索引（从0开始）
        :param color: QColor对象，背景色
        """
        # 校验行索引是否有效
        if row_index < 0 or row_index >= self.table.rowCount():
            print(f"行索引{row_index}无效！")
            return

        self.table.cellWidget(row_index, 0).setStyleSheet("QWidget { background-color: %s; }" % f"#{color.rgb():06X}")

        # 遍历该行的所有列，逐个设置单元格背景色
        for col in range(self.table.columnCount()):
            # 获取单元格项，如果不存在则创建（避免空单元格设置无效）
            item = self.table.item(row_index, col)
            if not item:
                item = QTableWidgetItem()
                self.table.setItem(row_index, col, item)
            # 设置背景色（QBrush用于绘制背景）
            item.setBackground(color)

    def clean_all_background(self):
        """
        清空表格所有行背景色
        """
        for row in range(self.table.rowCount()):
            self.set_row_background(row, QColor(255, 255, 255))

    def field_mapping(self, field_name: str, value) -> str:
        ret = ""
        if field_mapping := self.get_field_mapping():
            if kv := field_mapping.get(field_name):
                ret = kv.get(value)
        return ret
        # return self.get_field_mapping().get(field_name).get(value)

    def get_field_mapping(self) -> Dict:
        """
        获取字段名称映射
        数据库中存储状态值但是表格展示的时候需要展示状态名称
        :return:
        返回示例：{"upload_status": {0: "未传", 1: "成功", -1: "失败"}}
        "upload_status"字段为字段名称，0,1,-1为数据库中存储的内容，对应的中文为value
        """
        return {}

    def update_pagination(self, total_count: int):
        max_page = (total_count + self.page_size - 1) // self.page_size
        self.btn_prev.setEnabled(self.page_num > 1)
        self.btn_next.setEnabled(self.page_num < max_page)

    def update_selection(self, state: int, row: int):
        """更新选中行"""
        if state == Qt.Checked:
            self.selected_rows.add(row)
        else:
            self.selected_rows.discard(row)

    def get_selected_ids(self) -> List[int]:
        """获取选中行的ID"""
        ids = []
        for row in self.selected_rows:
            item = self.table.item(row, 1)  # 假设ID在第1列
            if item:
                ids.append(int(item.text()))
        return ids

    def build_query_condition(self) -> dict:
        """
        构建查询条件字典
        """
        condition = {}
        if self.is_need_search:
            for field, widget in self.query_widgets.items():
                if '_start' in field or '_end' in field:  # 范围字段，比如：日期
                    base_field = field.rsplit('_', 1)[0]
                    date_value = self._validate_date(widget.text().strip())

                    if date_value:
                        # 自动合并为范围条件
                        if field.endswith('_start'):
                            condition[f"{base_field}__gte"] = date_value
                        elif field.endswith('_end'):
                            condition[f"{base_field}__lte"] = date_value
                else:
                    value = None
                    if isinstance(widget, QComboBox):
                        value = widget.currentData()
                    elif isinstance(widget, QLineEdit):
                        value = widget.text().strip() or None
                    if value is not None:
                        condition[field] = value
        return condition

    def _validate_date(self, date_str: str) -> Optional[date]:
        """日期格式验证"""
        if not date_str:
            return None

        try:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            QMessageBox.warning(self, "格式错误", "日期格式应为YYYY-MM-DD")
            return None

    def create_pagination_ui(self):
        """创建分页组件"""
        pagination_layout = QHBoxLayout()

        self.lbl_total = QLabel("共0条")
        self.btn_prev = QPushButton("上一页")
        # 设置按钮的宽度
        self._optimize_button_width(self.btn_prev, "上一页")
        self.btn_prev.clicked.connect(self.prev_page)
        self.lbl_page = QLabel("1/1")
        self.btn_next = QPushButton("下一页")
        # 设置按钮的宽度
        self._optimize_button_width(self.btn_next, "下一页")
        self.btn_next.clicked.connect(self.next_page)
        self.le_page_num = QLineEdit()
        self.le_page_num.setMaximumWidth(120)
        self.le_page_num.setPlaceholderText("页码")
        self.le_page_num.textChanged.connect(self.validate_page_num)
        validator = QIntValidator(bottom=1)
        self.le_page_num.setValidator(validator)
        self.btn_go = QPushButton("Go")
        # 设置按钮的宽度
        self._optimize_button_width(self.btn_go, "Go")
        self.btn_go.clicked.connect(self.go_to_page)

        self.cmb_page_size = QComboBox()
        self.cmb_page_size.addItems(["10", "20", "30", "50", "100"])
        self.cmb_page_size.currentTextChanged.connect(self.change_page_size)

        pagination_layout.addWidget(self.lbl_total)
        pagination_layout.addWidget(QLabel("每页:"))
        pagination_layout.addWidget(self.cmb_page_size)
        pagination_layout.addWidget(self.btn_prev)
        pagination_layout.addWidget(self.lbl_page)
        pagination_layout.addWidget(self.btn_next)
        pagination_layout.addWidget(self.le_page_num)
        pagination_layout.addWidget(self.btn_go)
        pagination_layout.addStretch()

        return pagination_layout

    def _optimize_button_width(self, button, text):
        # 获取按钮字体的字体度量对象
        font_metrics = QFontMetrics(button.font())
        # 计算文本的宽度
        text_width = font_metrics.width(text)
        # 给文本宽度加上一些额外的空间，以确保按钮外观美观
        extra_space = 20
        # 设置按钮的宽度
        button.setFixedWidth(text_width + extra_space)
        # return text_width + extra_space

    @contextmanager
    def _show_progress(self, message: str):
        """进度提示上下文管理器"""
        progress = QProgressDialog(message, "取消", 0, 0, self)
        progress.setWindowModality(Qt.WindowModal)
        try:
            progress.show()
            QApplication.processEvents()  # 强制UI刷新
            yield
        finally:
            progress.close()

    def clear_all(self):
        reply = QMessageBox.question(
            self, "确认删除",
            f"确定要删除所有数据吗？",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            self.do_delete_all()

    def batch_delete(self):
        """批量删除操作"""
        selected_ids = self.get_selected_ids()
        if not selected_ids:
            QMessageBox.warning(self, "警告", "请先选择要删除的记录")
            return

        reply = QMessageBox.question(
            self, "确认删除",
            f"确定要删除选中的{len(selected_ids)}条记录吗？",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            if self.prepare_for_delete(selected_ids):
                self.do_delete(selected_ids)

    def prev_page(self):
        if self.page_num > 1:
            self.page_num -= 1
            self.search_with_progress()

    def next_page(self):
        total = self._get_total_count()
        max_page = (total + self.page_size - 1) // self.page_size
        if self.page_num < max_page:
            self.page_num += 1
            self.search_with_progress()

    def _get_total_count(self) -> int:
        """获取当前查询条件下的总记录数"""
        condition = self.build_query_condition()
        _, total = self.get_records(condition, 1, 1)  # 仅获取总数
        return total

    def prepare_for_search(self, show_text):
        self.search_progress_dialog = QProgressDialog(show_text, "取消", 0, 0, self)
        self.search_progress_dialog.setWindowModality(Qt.WindowModal)
        self.search_progress_dialog.show()

    def on_search_finished(self, status: bool, msg: str, payloads):
        if self.search_progress_dialog and self.search_progress_dialog.isVisible():
            # 关闭进度对话框
            self.search_progress_dialog.close()
        # 渲染表格数据
        self.render_table(status, msg, payloads)

    def search_with_progress(self, progress_dialog_show_text="正在查询中..."):
        self.prepare_for_search(progress_dialog_show_text)
        # QApplication.processEvents()  # 强制刷新UI
        self.async_get_data(self.on_search_finished)

    def validate_page_num(self):
        if self.le_page_num.text() and int(self.le_page_num.text()) > 0:
            self.btn_go.setEnabled(True)
        else:
            self.btn_go.setEnabled(False)
            # page_num = int(self.le_page_num.text())
            # if page_num < 1:
            #     self.le_page_num.setText("1")
            # elif page_num > self._get_total_count() // self.page_size + 1:
            #     self.le_page_num.setText(str(self._get_total_count() // self.page_size + 1))

    def go_to_page(self):
        """定位到指定页"""
        self.page_num = int(
            self.le_page_num.text()) if self.le_page_num.text() and self.le_page_num.text().strip() and int(
            self.le_page_num.text()) <= self.total_pages else self.total_pages
        self.search_with_progress("正在跳转指定页...")

    def change_page_size(self, text: str):
        """修改每页显示数量"""
        self.page_size = int(text)
        self.page_num = 1  # 重置到第一页
        self.search_with_progress("正在切换每页数量...")

    # 创建基础菜单（子类可重写此方法扩展）
    def create_context_menu(self):
        menu = QMenu(self)
        # 样式1：针对objectName为TaskMenu的QMenu（精准单独设置）
        """
            /* QMenu基础样式 */
            QMenu {
                background-color: #2c3e50;  /* 菜单背景色 */
                color: white;               /* 文字颜色 */
                font-size: 13px;            /* 字体大小 */
                border: 1px solid #34495e;  /* 边框 */
                padding: 5px 0;             /* 内边距（上下） */
            }
            
            /* 菜单项基础样式 */
            QMenu::item {
                height: 35px;               /* 菜单项高度 */
                padding: 0 20px;            /* 菜单项内边距（左右） */
            }
        """

        self.set_menu_style(menu)
        # 父类的基础菜单选项
        self.edit_action = menu.addAction("编辑")
        return menu

    def set_menu_style(self, menu):
        menu.setStyleSheet("""
            /* 菜单项选中/悬浮样式 */
            QMenu::item:selected {
                background-color: #198754;  /* 选中背景色 */
                color: white;
            }
            /* 分隔线样式 */
            QMenu::separator {
                height: 1px;                /* 分隔线高度 */
                background-color: #34495e;  /* 分隔线颜色 */
                margin: 5px 10px;           /* 分隔线边距 */
            }
        """)

        return menu

    def show_context_menu(self, pos):
        row = self.table.rowAt(pos.y())
        if row >= 0:
            menu = self.create_context_menu()
            # edit_action = menu.addAction("编辑")
            action = menu.exec_(self.table.viewport().mapToGlobal(pos))
            # 处理父类的菜单动作
            self.handle_menu_action(action, row)

    # 处理菜单动作（子类可重写此方法处理自定义动作）
    def handle_menu_action(self, action, row):
        if action == self.edit_action:
            self.edit_record(row)

    def get_selected_rows(self) -> List[Dict]:
        """
        获取选择的行的数据
        返回一行的数据
        :return:
        """
        return [self._get_row_data(row) for row in self.selected_rows]

    def get_all_rows(self):
        """获取指定行的原始数据"""
        return [self._get_row_data(row) for row in range(self.table.rowCount())]

    def _get_row_data(self, row: int) -> dict:
        """
        获取指定行的原始数据
        """
        data = {}
        headers = self.get_headers()  # 排除复选框列

        for col in range(1, self.table.columnCount()):  # 从第1列开始
            header_text = self.table.horizontalHeaderItem(col).text()
            field_name = next((h.field_name for h in headers if h.show_name == header_text), None)
            if field_name:
                item = self.table.item(row, col)
                data[field_name] = item.text() if item else ""
        return data

    def edit_record(self, row: int):
        """重构后的编辑方法：先请求最新元数据，再渲染弹窗"""
        # 1. 保存待编辑的行数据（基础数据，非下拉选项元数据）
        self.pending_edit_row = row  # 新增属性：待编辑行号
        self.pending_edit_data = self._get_row_data(row)  # 新增属性：待编辑行数据

        if not self.pending_edit_data:
            return

        # 2. 显示加载提示
        # self.edit_loading_dialog = QProgressDialog("正在加载最新数据...", "取消", 0, 0, self)
        # self.edit_loading_dialog.setWindowModality(Qt.WindowModal)
        # self.edit_loading_dialog.show()

        # 3. 异步请求最新的编辑元数据（如下拉选项）
        self.async_task_scheduler.submit_task(
            self.get_edit_metadata,  # 子类实现：获取最新的下拉选项等元数据
            finished_callback=self.on_edit_metadata_loaded
        )

    def on_edit_metadata_loaded(self, status: bool, msg: str, metadata: dict):
        """编辑元数据加载完成后的回调"""
        # 关闭加载提示
        # if self.edit_loading_dialog and self.edit_loading_dialog.isVisible():
        #     self.edit_loading_dialog.close()
        #     self.edit_loading_dialog = None

        if not status:
            QMessageBox.warning(self, "错误", f"加载最新数据失败：{msg}")
            return

        # 渲染编辑弹窗（结合最新元数据和待编辑行数据）
        self.render_edit_dialog(self.pending_edit_row, self.pending_edit_data, metadata)

    def render_edit_dialog(self, row: int, row_data: dict, metadata: dict):
        """渲染编辑弹窗（使用最新元数据）"""
        # 创建编辑对话框
        self.edit_dialog = QDialog(self)
        self.edit_dialog.setWindowTitle("编辑记录")
        self.edit_dialog.setMinimumWidth(300)
        layout = QVBoxLayout()
        # 表单布局
        form_layout = QFormLayout()
        # 第一列为ID列
        id_column = self.get_headers()[0]
        field_value = str(row_data.get(id_column.field_name, ""))
        form_layout.addRow(QLabel(id_column.show_name + ":"), QLabel(field_value))

        headers = self.get_headers()[1:]  # 跳过ID列

        for header in headers:
            if not header.is_edit_visible:
                continue
            field_name = header.field_name
            label_text = header.show_name
            field_value = str(row_data.get(field_name, ""))

            # 根据字段类型+最新元数据创建对应组件
            field = self._create_edit_widget_with_metadata(field_name, field_value, metadata)
            if field and field.widget:
                if not field.visible:
                    continue
                widget = field.widget
                if hasattr(widget, "setVisible"):
                    widget.setVisible(field.visible)
                if hasattr(widget, "setReadOnly"):
                    widget.setReadOnly(field.readonly)
                self.edit_fields[field_name] = field
                # if header.is_column_visible:
                form_layout.addRow(QLabel(label_text + ":"), field.widget)

        # 按钮布局（原有逻辑）
        button_layout = QHBoxLayout()
        btn_save = QPushButton("保存")
        btn_save.clicked.connect(lambda: self.save_edited_data(row))
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.edit_dialog.reject)

        button_layout.addWidget(btn_save)
        button_layout.addWidget(btn_cancel)

        layout.addLayout(form_layout)
        layout.addLayout(button_layout)
        self.edit_dialog.setLayout(layout)
        self.edit_dialog.exec_()

    def _create_edit_widget_with_metadata(self, field_name: str, value: str, metadata: dict) -> Optional[EditableField]:
        """根据字段类型+最新元数据创建编辑组件"""
        # 从元数据中获取该字段的最新属性（如下拉选项）
        field_attr = metadata.get(field_name, EditableField(field_name, "text", widget=QLineEdit()))

        field_type = field_attr.field_type
        widget = None
        if field_type == "select":
            widget = QComboBox()
            # 使用最新的下拉选项（来自元数据）
            options = field_attr.select_options
            for option in options:
                widget.addItem(option[0], option[1])
            # 匹配当前值（用实际值而非显示文本）
            current_index = next((i for i, (k, _) in enumerate(options) if str(k) == value), 0)
            widget.setCurrentIndex(current_index)
        elif field_type == "date":
            widget = QLineEdit(value)
            widget.setPlaceholderText("YYYY-MM-DD")
            widget.mousePressEvent = self.create_date_handler(widget)
        # elif field_type == "readonly":
        #     widget = QLabel(value)
        elif field_type == "text":
            widget = QLineEdit(value)
        elif field_type == "textedit":
            widget = QTextEdit(value)
            widget.setFixedHeight(120)
        elif field_type == "label":
            widget = QLabel(value)
        else:
            widget = None

        field_attr.widget = widget
        return field_attr

    # ---------------------- 新增抽象方法：子类实现获取最新元数据 ----------------------
    def get_edit_metadata(self) -> dict:
        """
        获取编辑所需的最新元数据（子类实现）
        返回：元数据字典
        元数据格式：{字段名: EditableFieldAttribute对象}
        """
        return {}

    def get_add_metadata(self) -> dict:
        """
        获取新增所需的最新元数据（子类实现）
        返回：元数据字典
        """
        return {}

    def save_edited_data(self, row: int):
        """保存编辑后的数据"""
        # 收集表单数据
        update_data = {}
        # headers = self.get_headers()[1:]
        # for header in headers:
        #     field_name = header[1]
        # field = self.edit_fields.get(field_name)
        # if not field or not field.widget:
        #     continue
        for field_name, field in self.edit_fields.items():
            widget = field.widget
            if isinstance(widget, QComboBox):
                # 获取字段属性配置
                if not field.related_field:
                    update_data[field_name] = widget.currentData()
                else:
                    update_data[field_name] = widget.currentText()
                    update_data[field.related_field] = widget.currentData()
            elif isinstance(widget, QLineEdit):
                update_data[field_name] = widget.text()
            elif isinstance(widget, QTextEdit):
                update_data[field_name] = widget.toPlainText()
            elif isinstance(widget, (QLabel, None)):
                continue  # 跳过只读字段

        # 执行更新
        record_id = int(self.table.item(row, 1).text())  # 假设ID在第1列
        if self.prepare_for_update(record_id, update_data):
            self.do_update(record_id, update_data)

    def prepare_for_update(self, record_id, update_data) -> bool:
        """
        更新的前置操作
        子类自行实现逻辑
        :param record_id: 记录ID
        :param update_data: 更新的内容
        :return: True-成功；False-失败
        """
        return True

    def prepare_for_add(self, add_data) -> bool:
        """
        添加的前置操作
        子类自行实现逻辑
        :param add_data: 添加的内容
        :return: True-成功；False-失败
        """
        return True

    def prepare_for_delete(self, record_ids: List[int]) -> bool:
        """
        删除的前置操作
        子类自行实现逻辑
        :param record_ids: 记录ID列表
        :return: True-成功；False-失败
        """
        return True

    def get_field_attribute(self, field_name: str) -> str:
        """
        编辑时（添加、编辑）获取字段属性（需子类实现）
        可选属性：
        text：文本（默认）
        select：下拉选择器
        date：日期
        hide：隐藏
        """
        return 'text'

    def get_add_field_attributes(self, field_name: str) -> List[EditableField]:
        return []

    def get_edit_field_attributes(self, field_name: str) -> List[EditableField]:
        return []

    # -------------------------- 辅助方法：简化关联字段配置（子类调用） --------------------------
    def create_related_value_getter(self, options: List[Tuple[str, Any]]) -> Callable[[int], Any]:
        """
        创建关联值获取器（通用方法，子类可直接调用）
        :param options: 下拉选项列表 [(显示文本, 实际值)]
        :return: 接收索引返回显示文本的函数
        """

        def getter(index: int) -> str:
            if 0 <= index < len(options):
                return options[index][0]  # 返回显示文本（如project_name）
            return ""

        return getter

    def get_field_options(self, field_name: str) -> List[Tuple[str, Any]]:
        """返回选择字段的选项列表（显示文本，实际值）"""
        return []

    def validate_import_data(self, data: dict) -> bool:
        """验证导入数据（子类实现）"""
        return True

    def add_one(self, data: dict):
        """
        新增单条数据（子类实现）
        若是同步方式新增数据，则需要返回：Tuple[bool, str, object] (success, msg, payloads)
        """
        self.async_task_scheduler.submit_task(self.get_add_one_callable(), data,
                                              finished_callback=self.on_add_one_result)

    @abstractmethod
    def get_add_one_callable(self) -> Callable:
        """
        获取添加数据的可调用方法
        返回内容：Callable类型，例如：一个方法名，
        方法示例：func(param1: dict)
        该方法需支持接收一个Dict类型的参数，为要添加的数据对象！将该对象持久化！
        :return:
        """
        pass

    def do_add_one(self, data: dict):
        """
        新增单条数据，对外暴露的统一新增入口（基类控制流程）
        自动处理同步/异步的新增结果，并触发后续操作
        """
        try:
            # 调用子类实现的更新方法
            result = self.add_one(data)

            # 同步操作：result有返回值，直接处理结果
            if result is not None:
                success, msg, payloads = result
                self.on_add_one_result(success, msg, payloads)
            # 异步操作：子类会发射 add_one_finished 信号，无需处理

        except Exception as e:
            # 统一捕获更新异常，触发失败逻辑
            self.on_add_one_result(False, f"新增失败：{str(e)}", {})

    def on_add_one_result(self, success: bool, msg: str, payloads):
        """
        数据新增后的统一处理逻辑
        add_one_finished信号的连接方法，子类可按需重写该方法，自定义新增记录后的操作
        如刷新表格、弹窗提示、日志记录等
        :param success: 状态；True-成功；False-失败
        :param msg: 消息
        :param payloads:  add_one方法的返回值
        :return:
        """
        if not success:
            QMessageBox.warning(self, "错误", msg)
        else:
            self.async_refresh_table()
            QMessageBox.information(self, "成功", "记录添加成功")
            # self.add_one_dialog.accept()

    def batch_insert(self, data_list: List[dict]):
        """批量插入数据（子类实现）"""
        self.async_task_scheduler.submit_task(self.get_batch_insert_callable(), data_list,
                                              finished_callback=self.on_batch_insert_result)

    def get_batch_insert_callable(self) -> Callable:
        pass

    def do_batch_insert(self, data_list: List[dict]):
        """
        对外暴露的统一批量插入口（基类控制流程）
        自动处理同步/异步的更新结果，并触发后续操作
        :param data_list: 批量插入的记录列表
        :return:
        """
        try:
            # 调用子类实现的更新方法
            result = self.batch_insert(data_list)
            # 同步操作：result有返回值，直接处理结果
            if result is not None:
                success, msg, payloads = result
                self.on_batch_insert_result(success, msg, payloads)
            # 异步操作：子类会发射 update_finished 信号，无需处理
        except Exception as e:
            # 统一捕获更新异常，触发失败逻辑
            self.on_batch_insert_result(False, f"更新失败：{str(e)}", {})

    def on_batch_insert_result(self, success: bool, msg: str, payloads):
        """
        数据插入后的统一处理逻辑
        update_finished信号的连接方法，子类可按需重写该方法，自定义更新后的操作
        如刷新表格、弹窗提示、日志记录等
        :param success: 状态；True-成功；False-失败
        :param msg: 消息
        :param payloads: 调用get_batch_insert_callable中返回的方法的返回值
        """
        if success:
            self.async_refresh_table()
            QMessageBox.information(self, "成功", "导入完成")
        else:
            QMessageBox.critical(self, "失败", f"数据入库失败：{msg}")

    def get_headers(self) -> List[TableHeader]:
        """返回表头配置（显示文本，字段名）"""
        return []

    def get_query_fields(self) -> List[QueryField]:
        """
        返回查询条件配置
        元素类型查看：QueryField
        """
        return []

    def get_records(self, condition: dict, page=1, page_size=0) -> Tuple[List[dict], int]:
        """
        获取分页记录
        返回：(分页数据, 总记录数)

        区间查询字段的值获取：
        支持区间查询的字段的上下限字段为在原先字段的名称后加上“__gte”和“__lte”
        比如create_date支持区间查询，则create_date区间下限的参数为：create_date__gte，上限参数为：create_date__lte。
        开发时从condition中获取create_date__gte和create_date__lte值，作为查询条件！
        :param condition: dict 查询条件
        :param page: int 页码
        :param page_size: int 页大小，=0时表示查询全部
        :return: Tuple[List[dict], int] (分类页数据列表, 总条数)
        """
        return [], 0

    def update_data(self, record_id: int, data: dict):
        """
        更新单条数据（子类实现）
        若是同步方式更新数据，则需要返回：Tuple[bool, str, object] (success, msg, payloads)
        :param record_id: 记录ID
        :param data: 更新的数据
        """
        self.async_task_scheduler.submit_task(self.get_update_callable(), record_id, data,
                                              finished_callback=self.on_update_result)

    def do_update(self, record_id: int, data: dict):
        """
        对外暴露的统一更新入口（基类控制流程）
        自动处理同步/异步的更新结果，并触发后续操作
        :param record_id: 记录ID
        :param data: 更新的内容。例如：{"name": "小明", "gender": "女"}
        :return:
        """
        try:
            # 调用子类实现的更新方法
            result = self.update_data(record_id, data)
            # 同步操作：result有返回值，直接处理结果
            if result is not None:
                success, msg, payloads = result
                self.on_update_result(success, msg, payloads)
            # 异步操作：子类会发射 update_finished 信号，无需处理
        except Exception as e:
            # 统一捕获更新异常，触发失败逻辑
            self.on_update_result(False, f"更新失败：{str(e)}", {})

    def on_update_result(self, success: bool, msg: str, payloads):
        """
        数据更新后的统一处理逻辑
        update_finished信号的连接方法，子类可按需重写该方法，自定义更新后的操作
        如刷新表格、弹窗提示、日志记录等
        :param success: 状态；True-成功；False-失败
        :param msg: 消息
        :param payloads: 调用get_update_callable返回的方法的返回值
        """
        if success:
            self.edit_dialog.accept()
            self.async_refresh_table()
            QMessageBox.information(self, "提示", "数据修改成功！")
        else:
            QMessageBox.critical(self, "警告", f"保存失败！错误：{msg}")

    @abstractmethod
    def get_update_callable(self) -> Callable:
        """
        获取更新数据的可调用方法
        返回内容：Callable类型，例如：一个方法名，
        方法示例：func(int, dict)
        第一个参数为记录ID，第二个参数为要更新的字段映射内容，例如：{"name": "小明", "gender": "男"}
        :return:
        """
        pass

    def delete_data(self, row_ids: List[int]):
        """
        删除多条数据（子类实现）
        若是同步方式删除多条数据，则需要返回：Tuple[bool, str, object] (success, msg, payloads)
        :param row_ids: 数据表中的记录ID，为第一列
        """
        self.async_task_scheduler.submit_task(self.get_delete_data_callable(), row_ids,
                                              finished_callback=self.on_delete_result)

    def do_delete(self, row_ids: List[int]):
        """
        对外暴露的统一删除入口（基类控制流程）
        自动处理同步/异步的删除结果，并触发后续操作
        :param row_ids: 记录ID列表
        :return:
        """
        try:
            # 调用子类实现的删除方法
            result = self.delete_data(row_ids)
            # 同步操作：result有返回值，直接处理结果
            if result is not None:
                success, msg, payloads = result
                self.on_delete_result(success, msg, payloads)
            # 异步操作：子类会发射 delete_finished 信号，无需处理
        except Exception as e:
            # 统一捕获更新异常，触发失败逻辑
            self.on_delete_result(False, f"删除失败：{str(e)}", {})

    def on_delete_result(self, success: bool, msg: str, payloads):
        """
        数据删除后的统一处理逻辑
        delete_finished/clear_all_finished信号的连接方法，子类可按需重写该方法，自定义更新后的操作
        如刷新表格、弹窗提示、日志记录等
        :param success: 状态；True-成功；False-失败
        :param msg: 消息
        :param payloads: 调用get_delete_data_callable返回的方法的返回值
        """
        if not success:
            QMessageBox.warning(self, "错误", msg)
        else:
            self.async_refresh_table()
            QMessageBox.information(self, "成功", "删除成功")

    @abstractmethod
    def get_delete_data_callable(self) -> Callable:
        """
        获取删除数据的可调用方法
        返回内容：Callable类型，例如：一个方法名，
        方法示例：func(List[int])
        第一个参数为要删除的记录ID列表
        :return:
        """
        pass

    def delete_all(self):
        """
        删除所有数据
        若是同步方式删除，则需要返回：Tuple[bool, str, object] (success, msg, payloads)
        """
        self.async_task_scheduler.submit_task(self.get_delete_all_callable(),
                                              finished_callback=self.on_delete_result)

    def get_delete_all_callable(self) -> Callable:
        """
        获取删除所有数据的可调用方法
        返回内容：Callable类型，例如：一个方法名，
        方法示例：func()
        :return:
        """
        pass

    def do_delete_all(self):
        """
        清空所有记录
        自动处理同步/异步的删除结果，并触发后续操作
        :return:
        """
        try:
            # 调用子类实现的删除方法
            result = self.delete_all()
            # 同步操作：result有返回值，直接处理结果
            if result is not None:
                success, msg, payloads = result
                self.on_delete_result(success, msg, payloads)
            # 异步操作：子类会发射 delete_finished 信号，无需处理
        except Exception as e:
            # 统一捕获更新异常，触发失败逻辑
            self.on_delete_result(False, f"清空失败：{str(e)}", {})

    def create_bulk_importer(self) -> BaseAsyncImportWorker:
        """创建导入工作线程（支持导入时必须实现）"""
        pass
